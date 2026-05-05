"""Generate Mindforge test notes Excel workbook. Run: python tools/generate_test_notes_workbook.py"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

OUT = Path(__file__).resolve().parent.parent / "docs" / "Mindforge_TestNotes.xlsx"

# Every row MUST include expected_similar_group: same string => should form one similar group in the app (threshold-dependent).

ROWS = [
    # ── Work / tech ──────────────────────────────────────────
    {
        "title": "Redis session cache – production rollout",
        "content": """We are enabling Redis for shared HTTP session storage across the three API instances behind the load balancer.
The cache TTL defaults to thirty minutes with sliding renewal on activity. Session keys are prefixed by tenant id to avoid collisions.
Rollout is staged: canary ten percent traffic tonight, full cutover Thursday if error budgets stay green.
Monitoring dashboards track hit ratio, eviction rate, and p95 session read latency.""",
        "search_queries": "shared cache for web sessions; scale-out session storage behind LB; staging a Redis cache launch",
        "group_theme": "Work — caching",
        "expected_similar_group": "W1-Redis-sessions",
        "merge_with_row": "2",
        "notes": "Near-duplicate with row 2; merge suggestion candidate.",
    },
    {
        "title": "Production Redis caching for user sessions",
        "content": """Rolling out Redis as our centralized store for web sessions spanning multiple API nodes behind the load balancer.
Default TTL is thirty minutes with refresh-on-use semantics. Keys namespaced by tenant identifier prevent cross-customer leakage.
Deployment plan: ten percent canary this evening, complete migration Thursday assuming SLOs hold.
We watch cache hit rate, evictions, and ninety-fifth percentile latency on session reads.""",
        "search_queries": "centralized session store Redis; multi-instance API session sharing; canary then full migration cache",
        "group_theme": "Work — caching",
        "expected_similar_group": "W1-Redis-sessions",
        "merge_with_row": "1",
        "notes": "Same group label as row 1 — should cluster.",
    },
    {
        "title": "Sprint standup – Mon Jan 6",
        "content": """Blocking: waiting on schema migration review from platform team. Yesterday finished the OAuth scope regression suite.
Today implementing rate-limit headers on the export endpoint. Carry-over story points look accurate if review lands by noon.
Team agreed to pair on flaky integration test after standup. No holiday conflicts this sprint.""",
        "search_queries": "daily scrum blockers OAuth; sprint planning carry-over points; flaky tests pairing after meeting",
        "group_theme": "Work — standup",
        "expected_similar_group": "W2-Standup-scrum",
        "merge_with_row": "4",
        "notes": "Near-duplicate with row 4.",
    },
    {
        "title": "Team standup notes Monday January sixth",
        "content": """Blockers: pending approval for database schema change from platform. Completed OAuth regression testing yesterday.
Focus today is adding rate limiting metadata to the bulk export API. Story points unchanged if we get review before lunch.
We'll mob on the unstable integration test right after standup. Everyone available for the full sprint.""",
        "search_queries": "schema approval blocker platform; bulk export rate limits; mob programming flaky test",
        "group_theme": "Work — standup",
        "expected_similar_group": "W2-Standup-scrum",
        "merge_with_row": "3",
        "notes": "Same group as row 3.",
    },
    {
        "title": "Coroutines on Android – best practices",
        "content": """Prefer viewModelScope for UI-related work tied to screen lifetime. Never block the main thread; offload IO with Dispatchers.IO.
Use suspend functions at repository boundaries and expose Flow to the UI layer when you need reactive streams.
Structured concurrency means child jobs cancel when the scope cancels, which prevents leaks after navigation.""",
        "search_queries": "viewModelScope lifecycle; suspend repository pattern; cancel coroutines when leaving screen",
        "group_theme": "Work — Android",
        "expected_similar_group": "W3-Android-coroutines",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "Structured concurrency with Kotlin coroutines",
        "content": """Coroutine scopes define when work must stop: activityScope deprecated patterns should move to lifecycle-aware APIs.
Combining async flows requires careful exception handling in supervisors versus strict failure propagation.
On Android, tying work to ViewModel scope aligns with configuration changes better than raw GlobalScope.""",
        "search_queries": "lifecycle-aware coroutine scopes; supervisor job exception handling; avoid GlobalScope mobile",
        "group_theme": "Work — Android",
        "expected_similar_group": "W3-Android-coroutines",
        "merge_with_row": "",
        "notes": "Clusters with row 5.",
    },
    {
        "title": "B-tree indexes for faster Postgres lookups",
        "content": """Equality and range filters on large tables benefit from btree indexes on selective columns. Composite indexes order matters: leading column should match the most selective predicate common in queries.
Avoid indexing low-cardinality booleans unless combined with other columns. Partial indexes shrink size when predicates are stable.""",
        "search_queries": "database index selectivity; composite column order SQL; partial index smaller footprint",
        "group_theme": "Work — database",
        "expected_similar_group": "W4-Postgres-tuning",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "PostgreSQL query planner and index selection",
        "content": """The planner estimates rows using statistics gathered by ANALYZE. Stale stats skew cost estimates and cause sequential scans on million-row tables.
When joins explode cardinality, consider join order hints only after validating with EXPLAIN. Prefer fixing statistics first.""",
        "search_queries": "ANALYZE stale statistics; sequential scan million rows; join cardinality explosion",
        "group_theme": "Work — database",
        "expected_similar_group": "W4-Postgres-tuning",
        "merge_with_row": "",
        "notes": "Clusters with row 7.",
    },
    {
        "title": "Docker Compose service discovery",
        "content": """Services reference each other by service name on the default bridge network. DNS resolution is handled by the embedded Docker resolver.
Override networks when isolating databases from frontend tiers. Healthchecks prevent dependents from starting before databases accept connections.""",
        "search_queries": "compose internal DNS service names; isolate DB network tier; healthcheck depends_on order",
        "group_theme": "Work — Compose",
        "expected_similar_group": "W5-Docker-compose",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "Networking between containers in Compose files",
        "content": """Published ports map host interfaces while internal ports stay cluster-private. Use aliases sparingly; default hostname equals service key.
For TLS termination, place reverse proxy in front of app containers and mount certificates read-only.""",
        "search_queries": "published ports versus internal; reverse proxy TLS compose; read-only cert mounts",
        "group_theme": "Work — Compose",
        "expected_similar_group": "W5-Docker-compose",
        "merge_with_row": "",
        "notes": "Clusters with row 9.",
    },
    {
        "title": "Semantic search with sentence embeddings",
        "content": """Dense vectors map queries and documents into shared space so paraphrases rank closer than keyword overlap alone.
Approximate nearest neighbor indexes trade recall for latency at scale. Re-ranking with a cross-encoder improves top-k quality.""",
        "search_queries": "dense vector retrieval paraphrase; ANN index latency recall; cross encoder rerank top results",
        "group_theme": "Work — ML",
        "expected_similar_group": "W6-ML-embeddings",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "ONNX Runtime on mobile devices",
        "content": """Mobile inference avoids round trips to servers when models fit memory budgets. Dynamic quantization pairs well with pruned attention layers.
Thread pools should stay off the UI thread; batch size one is typical for interactive similarity scoring.""",
        "search_queries": "on-device neural inference Android; quantized attention pruning; background thread inference UX",
        "group_theme": "Work — ML",
        "expected_similar_group": "W6-ML-embeddings",
        "merge_with_row": "",
        "notes": "Clusters with row 11.",
    },
    {
        "title": "Barcelona weekend itinerary draft",
        "content": """Saturday morning: Gothic Quarter walking tour then lunch near Born. Afternoon Sagrada Familia timed entry booked for four fifteen.
Sunday beach morning at Barceloneta, tapas crawl in Gracia evening. Hotel checkout Monday ten AM flight to Munich connection.""",
        "search_queries": "gotico walking route saturday; sagrada familia afternoon ticket; barceloneta sunday beach tapas gracia",
        "group_theme": "Travel",
        "expected_similar_group": "T1-Barcelona-trip",
        "merge_with_row": "14",
        "notes": "Near-duplicate with row 14.",
    },
    {
        "title": "Weekend trip plan Barcelona Spain",
        "content": """Day one explore Gothic Quarter walk followed by meal around El Born. Late afternoon entry at Sagrada Familia scheduled four fifteen slot.
Second day relax at Barceloneta beach morning, dinner crawl Gracia neighborhood. Monday morning checkout then flight via Munich.""",
        "search_queries": "el born lunch after old town; beach morning catalonia weekend; monday checkout flight connection",
        "group_theme": "Travel",
        "expected_similar_group": "T1-Barcelona-trip",
        "merge_with_row": "13",
        "notes": "Same trip as row 13.",
    },
    # ── Everyday — shopping ─────────────────────────────────
    {
        "title": "Saturday grocery run – Whole Foods",
        "content": """Pick up oatmeal, Greek yogurt, berries, spinach, cherry tomatoes, chicken thighs, olive oil if the big bottle is on sale.
Coffee beans medium roast—we are almost out. Grab whole milk for the kids and lactose-free for Alex.
Checkout line is brutal after eleven so aim for nine thirty. Reusable bags in the trunk remember.""",
        "search_queries": "supermarket yogurt berries chicken; milk and coffee beans restock; go early beat checkout line",
        "group_theme": "Life — shopping",
        "expected_similar_group": "L1-Groceries",
        "merge_with_row": "16",
        "notes": "Overlaps store-trip theme with 16–17; 15–16 strong merge candidate.",
    },
    {
        "title": "Grocery list before the week starts",
        "content": """Weekly staples we always burn through: oatmeal for breakfast bowls, berries and Greek yogurt, bagged spinach salad base.
Chicken thighs or tofu block for dinners, cherry tomatoes snackers, pantry needs olive oil if low. Whole milk toddlers, lactose-free carton for Alex.
Medium roast beans for drip coffee. Prefer Saturday Whole Foods nine thirty window when lines are sane. Toss six reusable bags in the car.""",
        "search_queries": "weekly staples yogurt oatmeal; tofu or chicken thighs dinner; lactose milk coffee beans saturday shop",
        "group_theme": "Life — shopping",
        "expected_similar_group": "L1-Groceries",
        "merge_with_row": "15",
        "notes": "Same groceries group label as rows 15, 17.",
    },
    {
        "title": "Costco bulk restock checklist",
        "content": """Not the weekly yogurt run—the big warehouse haul. Paper towels mega pack if garage shelf has space, dish pods, trash bags tall kitchen.
Frozen salmon fillets, mixed berries five pound bag, rotisserie chicken for quick lunches. Coffee two pound bag dark roast complements the drip grinder.
Parking is worst Sunday noon; weekdays after work lighter. Debit card renewed last month.""",
        "search_queries": "warehouse club paper towels dish pods; bulk frozen salmon berries; worst costco parking sunday noon",
        "group_theme": "Life — shopping",
        "expected_similar_group": "L1-Groceries",
        "merge_with_row": "",
        "notes": "Same similar group — food shopping wording.",
    },
    # ── Everyday — cooking ───────────────────────────────────
    {
        "title": "Sunday pasta sauce batch",
        "content": """Dice two onions caramelize low twenty minutes canned San Marzano crush by hand—not the immersion blender ruin texture.
Italian sausage browned first drain fat rosemary sprig tomato paste tablespoon cook Brick red basil finish off heat Parmesan rind simmer.
Makes eight portions freeze flat in quart freezer bags ninety minutes total hands on maybe twenty five.""",
        "search_queries": "marzano tomatoes slow simmer; sausage ragu freeze portions; caramelize onion tomato paste",
        "group_theme": "Life — cooking",
        "expected_similar_group": "L2-Cooking-meals",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "Meal prep – taco night ingredients",
        "content": """Ground beef or swap plant crumble if vegan guest. Corn tortillas char quick gas flame small bowl guacamole three avocados lime.
Black beans from can rinsed pico de gallo diced onion tomato cilantro fridge jar salsa verde sharper kids prefer mild cheddar shredded.
Prep bowls Wednesday night leftovers Thursday tacos again or burrito.""",
        "search_queries": "weekly taco ingredients ground beef; char corn tortillas avocado; salsa verde mild cheddar kids",
        "group_theme": "Life — cooking",
        "expected_similar_group": "L2-Cooking-meals",
        "merge_with_row": "",
        "notes": "Same cooking cluster as 18, 20.",
    },
    {
        "title": "Slow cooker veggie soup recipe",
        "content": """Chop celery carrots russet potato small dice toss crock eight hours low bay leaf thyme dried barley half cup rinsed.
Kidney beans can drained diced tomatoes quart vegetable broth carton low sodium skim fat layer if chilled overnight garnish parsley lemon squeeze.
Freeze two lunch portions weekday backup when nobody wants stove duty.""",
        "search_queries": "crock pot vegetable barley soup; low sodium broth bay leaf thyme; freezer lunch weekday backup",
        "group_theme": "Life — cooking",
        "expected_similar_group": "L2-Cooking-meals",
        "merge_with_row": "",
        "notes": "",
    },
    # ── Everyday — todos / home errands ────────────────────
    {
        "title": "Weekend errands checklist — around the apartment",
        "content": """Water plants fiddle leaf and pothos living room needs cup each Saturday sink filter replace under counter twist model three if red light blinks.
Laundry two loads colors then whites gym bag finally. Take recycling down chute before seven noise rule vacuum hallway runner cat hair.
Call super about drip kitchen faucet left handle—photo already on phone thread.""",
        "search_queries": "water plants saturday apartment; laundry gym bag finally; kitchen faucet drip super photo",
        "group_theme": "Life — home todos",
        "expected_similar_group": "L3-Home-todos",
        "merge_with_row": "",
        "notes": "",
    },
    {
        "title": "Household to-dos before trip",
        "content": """Change HVAC filter closet return fifteen by twenty size write on calendar next swap March. Empty fridge perishables gift neighbor milk.
Run dishwasher last night leave cracked open no smell. Hold mail USPS online form three day pause water plants extra cup each instructions sticky living room.
Spare keys in lockbox code already texted pet sitter.""",
        "search_queries": "change air filter before vacation; hold mail usps three days; water plants pet sitter spare key lockbox",
        "group_theme": "Life — home todos",
        "expected_similar_group": "L3-Home-todos",
        "merge_with_row": "",
        "notes": "Same home-chores group as 21, 23.",
    },
    {
        "title": "Saturday morning clean-up punch list",
        "content": """Strip bed wash sheets hot dry extra fluff towels bathroom rack replace if mildew smell. Dust TV stand blinds quick pass microfiber.
Vacuum runner hallway pet hair same as last weekend cat sheds spring. Kitchen sink disposal lemon ice grind smell reset.
If super does not answer faucet email board cc property manager photo attachment.""",
        "search_queries": "wash sheets towels saturday; vacuum hallway cat hair; disposal lemon ice smell; property manager email faucet",
        "group_theme": "Life — home todos",
        "expected_similar_group": "L3-Home-todos",
        "merge_with_row": "",
        "notes": "",
    },
    # ── Outliers (unique group each) ─────────────────────────
    {
        "title": "Residential parking permit renewal",
        "content": """City portal opens renewals on the first of March. Need plate number, proof of address PDF under two megabytes, and resident sticker photo if transferring zones.
Fee increased to forty-eight dollars this year. Deadline end of month avoids overnight street penalty.""",
        "search_queries": "city hall vehicle sticker march deadline; proof of address upload portal; resident zone transfer fee",
        "group_theme": "Outlier — admin",
        "expected_similar_group": "X-Admin-only",
        "merge_with_row": "",
        "notes": "Should not fall into shopping or cooking groups.",
    },
    {
        "title": "Upright piano tuning appointment",
        "content": """Technician scheduled Thursday two PM. Piano last tuned eighteen months ago so pitch may need slight pitch raise before fine tuning.
Recommend humidity forty-two to fifty-two percent in that room. Quiet house for forty-five minutes after visit.""",
        "search_queries": "pitch raise before fine tuning; humidity range acoustic piano; technician visit quiet home",
        "group_theme": "Outlier — home",
        "expected_similar_group": "X-Piano-only",
        "merge_with_row": "",
        "notes": "Unique label — no similar group with other rows.",
    },
]


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes"

    headers = [
        "Row",
        "Title",
        "Content",
        "Semantic_search_try_these_queries",
        "Expected_group_theme",
        "Expected_similar_group",
        "Likely_merge_pair_row",
        "Testing_notes",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, row in enumerate(ROWS, start=1):
        ws.append(
            [
                i,
                row["title"],
                row["content"],
                row["search_queries"],
                row["group_theme"],
                row["expected_similar_group"],
                row["merge_with_row"],
                row["notes"],
            ]
        )

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 5,
        "B": 40,
        "C": 68,
        "D": 48,
        "E": 20,
        "F": 22,
        "G": 12,
        "H": 36,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30

    guide = wb.create_sheet("How_to_test", 1)
    n = len(ROWS)
    guide_cells = [
        (f"Mindforge — {n} synthetic notes (work + everyday life)", Font(bold=True, size=14)),
        ("", None),
        (
            "Column **Expected_similar_group**: every row that shares the **same** value should land in the **same** similar-note group "
            "in the app (subject to your similarity threshold). Different labels => different expected groups. "
            "Outliers use a unique label each (X-…).",
            None,
        ),
        ("", None),
        (
            "Import Title + Content into the app. Search: tap **Semantic** with phrases from column D. "
            "Groups screen: refresh after notes exist.",
            None,
        ),
        ("", None),
        ("Expected groups (by label):", Font(bold=True)),
        (
            "W1-Redis-sessions → rows 1–2 | W2-Standup-scrum → 3–4 | W3-Android-coroutines → 5–6 | W4-Postgres-tuning → 7–8 | "
            "W5-Docker-compose → 9–10 | W6-ML-embeddings → 11–12 | T1-Barcelona-trip → 13–14 | "
            "L1-Groceries → 15–17 | L2-Cooking-meals → 18–20 | L3-Home-todos → 21–23 | "
            "X-Admin-only → 24 | X-Piano-only → 25",
            None,
        ),
        ("", None),
        ("Merge suggestions (designed strong pairs):", Font(bold=True)),
        ("Rows 1↔2, 3↔4, 13↔14, 15↔16 (long bodies; paraphrased).", None),
        ("", None),
        ("Everyday vs work balance:", Font(bold=True)),
        (
            "Rows 1–12 work/tech, 13–14 travel, 15–23 shopping / cooking / home todos, 24–25 outliers.",
            None,
        ),
    ]
    row_idx = 1
    for text, font in guide_cells:
        cell = guide.cell(row=row_idx, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            cell.font = font
        row_idx += 1
    guide.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"Wrote {OUT} ({n} rows)")


if __name__ == "__main__":
    main()
